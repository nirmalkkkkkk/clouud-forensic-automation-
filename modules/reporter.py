"""
=============================================================
   CLOUD FORENSICS AUTOMATION - HEALTHCARE EDITION
   Report Generator: CSV + XLSX + PDF

=============================================================
"""
import os
import csv
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── ReportLab (PDF) ─────────────────────────────────────── #
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                 Table, TableStyle, HRFlowable)
from reportlab.lib.enums import TA_CENTER, TA_LEFT


# ════════════════════════════════════════════════════════════ #
#  CSV REPORT
# ════════════════════════════════════════════════════════════ #
FIELDNAMES = [
    "Case ID", "Investigator", "System Name",
    "File Name", "File Path", "File Type",
    "File Size (KB)", "Permissions", "Anomaly Flag",
    "Created Time", "Modified Time", "Accessed Time",
    "SHA256 Hash", "MD5 Hash", "Status",
    "Risk Level", "Risk Score", "Category",
    "Entropy", "Timestomping", "Evidence Collected On",
]


def write_csv_report(report_path: str, evidence_data: list) -> str:
    """Write CSV evidence report. Returns the CSV path."""
    os.makedirs(os.path.dirname(report_path), exist_ok=True)
    csv_path = report_path if report_path.endswith(".csv") else report_path.replace(".xlsx", ".csv")

    try:
        with open(csv_path, mode="w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
            writer.writeheader()
            for record in evidence_data:
                writer.writerow({k: record.get(k, "") for k in FIELDNAMES})
    except Exception as e:
        print(f"[CSV ERROR] {e}")

    return csv_path


# ════════════════════════════════════════════════════════════ #
#  XLSX REPORT
# ════════════════════════════════════════════════════════════ #
RISK_COLORS = {
    "Critical": "C00000",
    "High":     "FF6600",
    "Medium":   "FFC000",
    "Low":      "00B050",
}

HDR_FILL  = PatternFill("solid", fgColor="0D1B2A")
HDR_FONT  = Font(name="Arial", bold=True, color="00D4FF", size=10)
CTR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LFT_ALIGN = Alignment(horizontal="left",   vertical="center", wrap_text=True)
BORDER    = Border(
    left   = Side(style="thin", color="2A3A4A"),
    right  = Side(style="thin", color="2A3A4A"),
    top    = Side(style="thin", color="2A3A4A"),
    bottom = Side(style="thin", color="2A3A4A"),
)


def write_xlsx_report(report_path: str, evidence_data: list) -> str:
    """Write styled XLSX evidence report. Returns the XLSX path."""
    xlsx_path = report_path.replace(".csv", ".xlsx") if report_path.endswith(".csv") else report_path
    os.makedirs(os.path.dirname(xlsx_path), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Forensic Evidence"

    # ── Header row ───────────────────────────────────────── #
    for col, field in enumerate(FIELDNAMES, 1):
        cell = ws.cell(row=1, column=col, value=field)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = CTR_ALIGN
        cell.border    = BORDER
    ws.row_dimensions[1].height = 30

    # ── Data rows ────────────────────────────────────────── #
    alt_fill = PatternFill("solid", fgColor="0A1520")
    std_fill = PatternFill("solid", fgColor="0D1B2A")

    for row_num, record in enumerate(evidence_data, 2):
        row_fill = alt_fill if row_num % 2 == 0 else std_fill
        for col, field in enumerate(FIELDNAMES, 1):
            cell           = ws.cell(row=row_num, column=col, value=record.get(field, ""))
            cell.font      = Font(name="Arial", size=9, color="C8D8E8")
            cell.alignment = LFT_ALIGN
            cell.border    = BORDER
            cell.fill      = row_fill

        # Colour risk level cell
        risk_col  = FIELDNAMES.index("Risk Level") + 1
        risk_val  = record.get("Risk Level", "")
        if risk_val in RISK_COLORS:
            rc       = ws.cell(row=row_num, column=risk_col)
            rc.fill  = PatternFill("solid", fgColor=RISK_COLORS[risk_val])
            rc.font  = Font(name="Arial", bold=True, size=9, color="FFFFFF")
            rc.alignment = CTR_ALIGN

    # ── Column widths ────────────────────────────────────── #
    widths = {
        "Case ID": 14, "Investigator": 14, "System Name": 24,
        "File Name": 28, "File Path": 40, "File Type": 10,
        "File Size (KB)": 14, "Permissions": 26, "Anomaly Flag": 28,
        "Created Time": 20, "Modified Time": 20, "Accessed Time": 20,
        "SHA256 Hash": 66, "MD5 Hash": 36, "Status": 28,
        "Risk Level": 12, "Risk Score": 12, "Category": 28,
        "Entropy": 10, "Timestomping": 14, "Evidence Collected On": 22,
    }
    for col, field in enumerate(FIELDNAMES, 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(field, 18)

    ws.freeze_panes = "A2"

    # ── Summary sheet ────────────────────────────────────── #
    ws2 = wb.create_sheet(title="Summary")
    ws2["A1"] = "FORENSIC INVESTIGATION SUMMARY"
    ws2["A1"].font      = Font(name="Arial", bold=True, size=14, color="0D6EFD")
    ws2["A1"].alignment = CTR_ALIGN
    ws2.merge_cells("A1:B1")
    ws2.row_dimensions[1].height = 30

    summary = [
        ("Total Files Scanned",  len(evidence_data)),
        ("Critical Risk Files",  sum(1 for f in evidence_data if f.get("Risk Level") == "Critical")),
        ("High Risk Files",      sum(1 for f in evidence_data if f.get("Risk Level") == "High")),
        ("Medium Risk Files",    sum(1 for f in evidence_data if f.get("Risk Level") == "Medium")),
        ("Low Risk Files",       sum(1 for f in evidence_data if f.get("Risk Level") == "Low")),
        ("Anomalies Detected",   sum(1 for f in evidence_data if f.get("Anomaly Flag", "None") != "None")),
        ("Timestomping Flags",   sum(1 for f in evidence_data if f.get("Timestomping"))),
    ]
    for i, (label, value) in enumerate(summary, 2):
        ws2.cell(row=i, column=1, value=label).font  = Font(name="Arial", bold=True, size=10)
        ws2.cell(row=i, column=2, value=value).font  = Font(name="Arial", size=10)
        ws2.cell(row=i, column=1).border = BORDER
        ws2.cell(row=i, column=2).border = BORDER

    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 14

    wb.save(xlsx_path)
    return xlsx_path


# ════════════════════════════════════════════════════════════ #
#  PDF REPORT
# ════════════════════════════════════════════════════════════ #
_DARK_BG   = colors.HexColor("#0D1B2A")
_ACCENT    = colors.HexColor("#00D4FF")
_TEXT      = colors.HexColor("#C8D8E8")
_CRITICAL  = colors.HexColor("#C00000")
_HIGH      = colors.HexColor("#FF6600")
_MEDIUM    = colors.HexColor("#FFC000")
_LOW       = colors.HexColor("#00B050")

_RISK_COLOR_MAP = {
    "Critical": _CRITICAL,
    "High":     _HIGH,
    "Medium":   _MEDIUM,
    "Low":      _LOW,
}


def write_pdf_report(pdf_path: str, evidence_data: list,
                     case_id: str, investigator: str, scan_dir: str) -> str:
    """Generate a professional PDF forensic report. Returns the PDF path."""
    os.makedirs(os.path.dirname(pdf_path), exist_ok=True)

    doc = SimpleDocTemplate(
        pdf_path,
        pagesize=landscape(A4),
        rightMargin=1.5*cm, leftMargin=1.5*cm,
        topMargin=2*cm, bottomMargin=2*cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title2", parent=styles["Title"],
                                  textColor=_ACCENT, fontSize=22, alignment=TA_CENTER,
                                  spaceAfter=6)
    sub_style   = ParagraphStyle("Sub2",   parent=styles["Normal"],
                                  textColor=_TEXT,   fontSize=10, alignment=TA_CENTER,
                                  spaceAfter=4)
    label_style = ParagraphStyle("Label",  parent=styles["Normal"],
                                  textColor=_ACCENT, fontSize=9,  bold=True)
    body_style  = ParagraphStyle("Body",   parent=styles["Normal"],
                                  textColor=_TEXT,   fontSize=8)

    story = []

    # ── Cover ─────────────────────────────────────────────── #
    story.append(Paragraph("CLOUD FORENSICS AUTOMATION", title_style))
    story.append(Paragraph("Healthcare Data Breach Investigation Report", sub_style))
    story.append(HRFlowable(width="100%", thickness=1, color=_ACCENT, spaceAfter=10))

    meta = [
        ["Case ID:",       case_id,        "Investigator:", investigator],
        ["Scan Target:",   scan_dir,       "Generated:",    datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Total Files:",   len(evidence_data), "", ""],
    ]
    meta_table = Table(meta, colWidths=[3.5*cm, 7*cm, 3.5*cm, 7*cm])
    meta_table.setStyle(TableStyle([
        ("TEXTCOLOR",  (0, 0), (-1, -1), _TEXT),
        ("TEXTCOLOR",  (0, 0), (0, -1),  _ACCENT),
        ("TEXTCOLOR",  (2, 0), (2, -1),  _ACCENT),
        ("FONTNAME",   (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE",   (0, 0), (-1, -1), 9),
        ("FONTNAME",   (0, 0), (0, -1),  "Helvetica-Bold"),
        ("FONTNAME",   (2, 0), (2, -1),  "Helvetica-Bold"),
        ("BACKGROUND", (0, 0), (-1, -1), _DARK_BG),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.HexColor("#0A1520"), _DARK_BG]),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#1A3050")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 0.5*cm))

    # ── Risk Summary ──────────────────────────────────────── #
    story.append(Paragraph("RISK SUMMARY", label_style))
    story.append(HRFlowable(width="100%", thickness=0.5, color=_ACCENT, spaceAfter=6))

    counts = {lvl: sum(1 for f in evidence_data if f.get("Risk Level") == lvl)
              for lvl in ["Critical", "High", "Medium", "Low"]}
    anomalies  = sum(1 for f in evidence_data if f.get("Anomaly Flag", "None") != "None")
    timestomps = sum(1 for f in evidence_data if f.get("Timestomping"))

    summary_data = [
        ["Category", "Count", "% of Total"],
        ["Critical Risk", counts["Critical"], f"{100*counts['Critical']//max(len(evidence_data),1)}%"],
        ["High Risk",     counts["High"],     f"{100*counts['High']//max(len(evidence_data),1)}%"],
        ["Medium Risk",   counts["Medium"],   f"{100*counts['Medium']//max(len(evidence_data),1)}%"],
        ["Low Risk",      counts["Low"],       f"{100*counts['Low']//max(len(evidence_data),1)}%"],
        ["Anomalies",     anomalies,           f"{100*anomalies//max(len(evidence_data),1)}%"],
        ["Timestomping",  timestomps,          f"{100*timestomps//max(len(evidence_data),1)}%"],
    ]
    sum_table = Table(summary_data, colWidths=[6*cm, 3*cm, 3*cm])
    sum_styles = [
        ("BACKGROUND",    (0, 0), (-1, 0),  colors.HexColor("#0A2040")),
        ("TEXTCOLOR",     (0, 0), (-1, 0),  _ACCENT),
        ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, -1), 9),
        ("TEXTCOLOR",     (0, 1), (-1, -1), _TEXT),
        ("BACKGROUND",    (0, 1), (-1, -1), _DARK_BG),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.HexColor("#0A1520"), _DARK_BG]),
        ("GRID",          (0, 0), (-1, -1), 0.5, colors.HexColor("#1A3050")),
        ("ALIGN",         (1, 0), (-1, -1), "CENTER"),
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]
    # colour risk rows
    risk_row_map = {"Critical": 1, "High": 2, "Medium": 3, "Low": 4}
    for level, row in risk_row_map.items():
        sum_styles.append(("TEXTCOLOR", (0, row), (0, row), _RISK_COLOR_MAP[level]))
    sum_table.setStyle(TableStyle(sum_styles))
    story.append(sum_table)
    story.append(Spacer(1, 0.5*cm))

    # ── Evidence Table (top 60 by risk score) ────────────── #
    story.append(Paragraph("EVIDENCE DETAIL TABLE (Top 60 by Risk Score)", label_style))
    story.append(HRFlowable(width="100%", thickness=0.5, color=_ACCENT, spaceAfter=6))

    sorted_ev = sorted(evidence_data,
                       key=lambda x: int(x.get("Risk Score", 0) or 0),
                       reverse=True)[:60]

    ev_hdr = ["#", "File Name", "Type", "Size KB", "Risk Level", "Score",
              "Category", "Status", "Modified Time", "Anomaly"]
    ev_rows = [ev_hdr]
    for i, f in enumerate(sorted_ev, 1):
        ev_rows.append([
            str(i),
            str(f.get("File Name", ""))[:35],
            str(f.get("File Type", "")),
            str(f.get("File Size (KB)", "")),
            str(f.get("Risk Level", "")),
            str(f.get("Risk Score", "")),
            str(f.get("Category", ""))[:22],
            str(f.get("Status", ""))[:20],
            str(f.get("Modified Time", ""))[:19],
            str(f.get("Anomaly Flag", "None"))[:20],
        ])

    col_w = [0.8*cm, 5.5*cm, 1.4*cm, 1.8*cm, 2*cm, 1.4*cm, 4*cm, 4*cm, 3.8*cm, 3.8*cm]
    ev_table = Table(ev_rows, colWidths=col_w, repeatRows=1)
    ev_ts = [
        ("BACKGROUND",    (0, 0), (-1, 0),  colors.HexColor("#0A2040")),
        ("TEXTCOLOR",     (0, 0), (-1, 0),  _ACCENT),
        ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, -1), 7),
        ("FONTNAME",      (0, 1), (-1, -1), "Helvetica"),
        ("TEXTCOLOR",     (0, 1), (-1, -1), _TEXT),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.HexColor("#0A1520"), _DARK_BG]),
        ("GRID",          (0, 0), (-1, -1), 0.3, colors.HexColor("#1A3050")),
        ("ALIGN",         (0, 0), (0, -1),  "CENTER"),
        ("ALIGN",         (3, 0), (5, -1),  "CENTER"),
        ("TOPPADDING",    (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]
    # colour risk level cells
    for row_idx, row_data in enumerate(ev_rows[1:], 1):
        lvl = row_data[4]
        clr = _RISK_COLOR_MAP.get(lvl)
        if clr:
            ev_ts.append(("TEXTCOLOR", (4, row_idx), (4, row_idx), clr))
            ev_ts.append(("FONTNAME",  (4, row_idx), (4, row_idx), "Helvetica-Bold"))
    ev_table.setStyle(TableStyle(ev_ts))
    story.append(ev_table)

    doc.build(story)
    return pdf_path